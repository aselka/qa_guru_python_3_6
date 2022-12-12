import os
import zipfile
import pytest
import glob
import csv
from os.path import basename
from PyPDF2 import PdfReader
from openpyxl import load_workbook


path_to_files = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
path_resources = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
files_dir = os.listdir(path_to_files)
path_zip = os.path.join(path_resources, "test.zip")


@pytest.fixture()
def clear_dir():
    all_files = os.path.join(path_resources, '*.*')
    for file in glob.glob(all_files):
        os.remove(file)


def test_create_archive(clear_dir):
    with zipfile.ZipFile(path_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in files_dir:
            add_file = os.path.join(path_to_files, file)
            zf.write(add_file, basename(add_file))
    files = os.listdir(path_resources)
    assert len(files) == 1, f"Неверное количество скачанных файлов {len(files)}, не ровно {1}"
    assert "test.zip" == files[0], f"Архив {files[0]} создался с неправильным именем"


def test_pdf():
    with zipfile.ZipFile(path_zip) as myzip:
        pdf_f = myzip.extract("1.pdf")
        pdf_data = PdfReader(pdf_f)
        page = pdf_data.pages[0]
        text = page.extract_text()
        assert "Пример" in text
        os.remove(pdf_f)


def test_csv():
    with zipfile.ZipFile(path_zip) as myzip:
        csv_f = myzip.extract("2.csv")
        with open(csv_f) as csv_file:
            table = csv.reader(csv_file, delimiter=";")
            for line_no, line in enumerate(table, 1):
                if line_no == 2:
                    assert 'Иванова' in line[1]
        os.remove(csv_f)


def test_xlsx():
    with zipfile.ZipFile(path_zip) as myzip:
        xlsx_f = myzip.extract("3.xlsx")
        workbook = load_workbook(xlsx_f)
        sheet = workbook.active
        sheet = sheet.cell(row=3, column=2).value
        assert "Сергеев" in sheet
        os.remove(xlsx_f)


